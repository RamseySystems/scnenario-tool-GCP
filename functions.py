import shutil
import openpyxl
import csv
import json
import re
import html
import os
from jinja2 import Environment, FileSystemLoader


def clear_dir(folder_path: str):
    if not folder_path.startswith('/tmp'):
        raise ValueError(f"The folder '{folder_path}' does not start with /tmp.")
    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"The folder '{folder_path}' was created.")
    else:
        print(f"The folder '{folder_path}' already exists.")
        # Delete the contents of the folder
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            try:
                if os.path.isfile(file_path):
                    os.remove(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print(f"Error deleting {file_path}: {str(e)}")
        print(f"The contents of the folder '{folder_path}' were deleted.")

def render_template(template_name: str, data: dict, save_path: str):
    '''
    A function to render data against jinja templates

    :param template_name: the name of the template to use
    :param data: the data to pass to the template
    :param save_path: where to save it
    :return: :dict:
    '''
    env = Environment(loader=FileSystemLoader('templates'))
    template = env.get_template(template_name)
    output = template.render(data=data)

    with open(save_path, 'w') as f:
        f.write(output)

    return


def save_obj_to_file(obj, save_path):
    with open(save_path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=4, ensure_ascii=False)


def allowed_file(filename: str, allowed_extentions):
    return '.' in filename and \
        filename.rsplit('.', 1)[1].lower() in allowed_extentions


def process_file(file, save_dir, upload_folder, standard_path):
    # check for the type of file
    if file.split('.')[-1] == 'json':
        print('JSON')
        summary = {
            'file_path': f'{upload_folder}/{file}',
            'reformatted_json': '',
            'json_schema': ''
        }

        with open(f'{upload_folder}/{file}', 'r') as f:
            unformatted = json.load(f)

        reformatted_with_implamentation = filter(
            unformatted, implementationGuidance=True)
        reformatted = filter(unformatted, implementationGuidance=False)

        save_obj_to_file(reformatted_with_implamentation,
                         f'{save_dir}/{file[:-4]}_no_implamentation_reformatted.json')
        save_obj_to_file(
            reformatted, f'{save_dir}/{file[:-4]}_reformatted.json')

    elif file.split('.')[-1] == 'xlsx':

        # add summary info to summary object
        summary = {
            'personae': f'{file[:-5]}',
            'sheets': []
        }

        # create save dir for excel workbook
        os.mkdir(f'{save_dir}/{file[:-5]}')

        # create dir in website folder
        os.mkdir(f'{save_dir}/website/data/{file[:-5]}')
        os.mkdir(f'{save_dir}/website/data/{file[:-5]}/rendered')
        os.mkdir(f'{save_dir}/website/data/{file[:-5]}/json')

        # load each sheet in file intop iterable
        sheets, group_alias_sheet, path_alias_sheet = open_worksheets(
            f'{upload_folder}/{file}')

        # generate paths for standards
        standard_path_lists = []
        for standard in os.listdir(standard_path):
            current_standard_path = f'{standard_path}/{standard}'
            standard_paths = get_standard_paths(current_standard_path)
            standard_path_lists.append(standard_paths.copy())

        # loop over sheets
        for sheet in sheets:
            if sheet.title == 'Story':
                summary_cell = find_cell_location(sheet, 'Summary')
                rationale_cell = find_cell_location(sheet, 'Rationale')
                story_cell = find_cell_location(sheet, 'Story')
                standard_url_cell = find_cell_location(sheet, 'Standard URL')
                standard_name_cell = find_cell_location(sheet, 'Standard Name')

                story = {}

                story['summary'] = sheet.cell(
                    row=summary_cell[0]+1, column=summary_cell[1]).value
                story['rationale'] = sheet.cell(
                    row=rationale_cell[0]+1, column=rationale_cell[1]).value
                story['story'] = sheet.cell(
                    row=story_cell[0]+1, column=story_cell[1]).value
                story['standard_url'] = sheet.cell(
                    row=standard_url_cell[0]+1, column=standard_url_cell[1]).value
                story['standard_name'] = sheet.cell(
                    row=standard_name_cell[0]+1, column=standard_name_cell[1]).value

                file_save_path = f'{save_dir}/{file[:-5]}/story.json'

                with open(file_save_path, "w") as f:
                    json.dump(story, f, indent=4)

                template_name = 'story.html'
                html_save_path = f'{save_dir}/{file[:-5]}/{file[:-5]}_story.html'
                render_template(template_name, story, html_save_path)
                story_dir = f'{save_dir}/website/Stories'
                if not os.path.exists(story_dir):
                    os.makedirs(story_dir)
                html_save_path = f'{story_dir}/{file[:-5]}_story.html'
                render_template(template_name, story, html_save_path)

            elif sheet.title == 'Time Line':
                time_cell = find_cell_location(sheet, 'Date/Time')
                event_cell = find_cell_location(sheet, 'Event')
                sheet_cell = find_cell_location(sheet, 'Sheet')

                time_line = []

                for row in sheet.iter_rows(time_cell[0]+1, sheet.max_row, time_cell[1], sheet_cell[1]):
                    time = row[0].value
                    event = row[1].value
                    work_sheet = row[2].value

                    if str(time) != 'None':

                        time_line.append({
                            'time': time,
                            'event': event,
                            'sheet': work_sheet
                        })


                # render log 
                env = Environment(loader=FileSystemLoader('templates'))
                template = env.get_template('log.html')
                output = template.render(data=time_line, name=file[:-5])
                log_dir = f'{save_dir}/website/logs'
                if not os.path.exists(log_dir):
                    os.makedirs(log_dir)
                with open(f'{log_dir}/{file[:-5]}_log.html', 'w') as f:
                    f.write(output)


                file_save_path = f'{save_dir}/{file[:-5]}/time_line.json'

                with open(file_save_path, "w") as f:
                    json.dump(time_line, f, indent=4)

                template_name = 'timeline.html'
                html_save_path = f'{save_dir}/{file[:-5]}/time_line.html'

                render_template(template_name, time_line, html_save_path)
            else:

                # get path cell and value cell
                path_cell = find_cell_location(sheet, 'Data Path')
                value_cell = find_cell_location(sheet, 'Example Data')

                # generate path list
                path_list = gen_path_list(
                    sheet, path_cell, value_cell, group_alias_sheet, path_alias_sheet)

                # extract paths with no indexes
                plain_path_list = extract_paths(path_list, plain=True)
                combined_list = get_paths_and_values(path_list)

                # check if path list is valid
                invalid_paths = validate_path_list(
                    plain_path_list, standard_path_lists)

                # generate object
                object = create_object(combined_list)

                # save dir
                file_save_path = f'{save_dir}/{file[:-5]}/{sheet.title}.json'

                # save object to file
                with open(file_save_path, "w") as f:
                    json.dump(object, f, indent=4)

                # rendering the HTML file
                template_name = 'json_render.html'
                html_save_path = f'{save_dir}/{file[:-5]}/{sheet.title}.html'
                render_template(template_name, object, html_save_path)

                # save the files to website folder
                html_save_path = f'{save_dir}/website/data/{file[:-5]}/rendered/{sheet.title}.html'
                render_template(template_name, object, html_save_path)

                file_save_path = f'{save_dir}/website/data/{file[:-5]}/json/{sheet.title}.json'
                with open(file_save_path, "w") as f:
                    json.dump(object, f, indent=4)



                # add invalid paths to sheet summary
                summary['sheets'].append({
                    'sheet_name': sheet.title,
                    'invalid_paths': invalid_paths,
                    'object_save_path': file_save_path
                })

        return summary

    else:
        raise TypeError('Not a valid input file type: ' + file)


def open_worksheets(path: str):
    '''
    A function to load all the sheets from an excel file into an array containing each sheet as a object
    returns a python arr

    :param path: the path to the excel document
    :param sheet_name: the sheet to open
    :return: :dict:
    '''
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    wb_array = []
    [wb_array.append(wb[ws]) for ws in sheet_names if ws.lower()
     not in ['group aliases', 'path aliases']]
    try:
        group_alias_sheet = wb['Group Aliases']
    except ValueError:
        raise ValueError('Workbook needs a group alias sheet')

    try:
        path_alias_sheet = wb['Path Aliases']
    except ValueError:
        raise ValueError('Workbook needs a path alias sheet')

    return wb_array, group_alias_sheet, path_alias_sheet


def find_cell_location(sheet: dict,
                       cell_value: str):
    '''
    A function to return the coordinates to a cell in a worksheet
    Returns a tuple

    :param sheet: the sheet to search
    :param cell_value: the text to seatch for in the sheet
    :return: :tuple:
    '''
    row_index = 1
    for row in sheet.iter_rows():
        cell_index = 1
        for cell in row:
            value = str(cell.value).lower()
            if value == cell_value.lower():
                return (row_index, cell_index)
            cell_index += 1
        row_index += 1
    raise ValueError(
        f'\'{cell_value}\' does not exist in worksheet: {sheet.title}')


def translate_cell(cell_value: str):
    '''
    A function to expand and translate the case of the data in a cell
    Blank at the moment to discuss with charlie what the cells will be (will there be aliaes and what case will it be in)

    :param cell_value: the cell value to translate
    :return: :str:
    '''
    return cell_value.replace(' ', '_')


def gen_path_list(sheet: dict,
                  path_cell: tuple,
                  value_cell: tuple,
                  group_alias_sheet: dict,
                  path_alias_sheet: dict):
    '''
    A function that creates a path list array from the paths in a sheet. 
    Returns an array

    :param sheet: the sheet to get the paths from
    :param path_cell: the coordinates of the title cell for the paths
    :param group_alias_sheet: the sheet containing the expanded group aliases
    :param path_alias_sheet: the sheet containing the path aliases
    :return: :list:
    '''
    path_list = []
    in_loop = False
    for cell in sheet.iter_rows(path_cell[0]+1, sheet.max_row, path_cell[1], value_cell[1]):
        if str(cell[0].value) != 'None':

            # get cell value
            path = translate_cell(cell[0].value)
            value = str(cell[-1].value).replace('.', '')

            if in_loop:
                if path != '$loopend':
                    continue
                elif path == '$loopend':
                    in_loop = False
                    continue

            elif '$$' in path:

                # get the alias name
                group_alias_name = path.split(
                    '.')[-1].replace('$$', '').lower()

                # get the expanded list of paths
                group_alias_paths = expand_group_alias(
                    group_alias_name, group_alias_sheet)

                group_alias_paths_with_start = []
                no_group_alias = '.'.join(path.split('.')[:-1])

                for line in group_alias_paths:
                    joined_line = '.'.join(line)
                    group_alias_paths_with_start.append(
                        f'{no_group_alias}.{joined_line}')

                # look for path aliases
                for line in group_alias_paths_with_start:
                    split_path = line.split('.')
                    expanded_path = []
                    for element in split_path:
                        if '$' in element:
                            expanded_element = expand_path_alias(
                                path_alias_sheet, element)
                            expanded_path.append(expanded_element)
                        else:
                            expanded_path.append(element)

                    expanded_path_str = '.'.join(expanded_path)
                    path_list.append([expanded_path_str, value])

            elif '$loop' in path:

                in_loop = True

                # get csv name and open it
                csv_name = path.strip().split()[-1]
                csv_path = f'./continuous_data/{csv_name}'

                with open(csv_path, 'r') as f:
                    continuous_data = list(csv.reader(f))

                # get loop lines
                loop_start = cell[0].row
                loop_lines = get_loop_lines(
                    sheet, loop_start, path_cell, value_cell)
                expanded_loop_lines = expand_loop_lines(
                    loop_lines, continuous_data, path_alias_sheet, group_alias_sheet)

                [path_list.append(line) for line in expanded_loop_lines]

            elif '$' in path:

                # get alias
                split_path = path.split('.')
                expanded_path = []
                for element in split_path:
                    if '$' in element:
                        expanded_element = expand_path_alias(
                            path_alias_sheet, element)
                        expanded_path.append(expanded_element)
                    else:
                        expanded_path.append(element)

                expanded_path_str = '.'.join(expanded_path)
                path_list.append([expanded_path_str, value])

            else:
                path_list.append([path, value])

    return path_list


def gen_path_list_new(sheet: dict,
                      path_cell: tuple,
                      value_cell: tuple,
                      group_alias_sheet: dict,
                      path_alias_sheet: dict):
    '''
    A function that creates a path list array from the paths in a sheet. 
    Returns an array

    :param sheet: the sheet to get the paths from
    :param path_cell: the coordinates of the title cell for the paths
    :param group_alias_sheet: the sheet containing the expanded group aliases
    :param path_alias_sheet: the sheet containing the path aliases
    :return: :list:
    '''
    path_list = []
    in_loop = False
    for cell in sheet.iter_rows(path_cell[0]+1, sheet.max_row, path_cell[1], value_cell[1]):
        if str(cell[0].value) != 'None':

            # get cell value
            path = translate_cell(cell[0].value)
            value = cell[-1].value

            if in_loop:
                if path != '$loopend':
                    continue
                elif path == '$loopend':
                    in_loop = False
                    continue

            # take apart path into sections and replace group alias
            path_split = path.split('.')
            built_path = []
            for section in path_split:
                if '$$' in section:

                    # get the alias name
                    group_alias_name = section.replace(
                        '$$', '').strip().lower()

                    # get the expanded list of paths
                    group_alias_paths = expand_group_alias(
                        group_alias_name, group_alias_sheet)

                    # loop over list and add each line item to path list
                    for path in group_alias_paths:
                        if len(built_path) == 0:
                            path_list.append(path)
                        else:
                            built_path_str = '.'.join(built_path)
                            path_list.append(f'{built_path_str}.{path}')

                elif '$' in section:

                    # get alias
                    split_path = section.split('.')
                    expanded_path = []
                    for element in split_path:
                        if '$' in element:
                            expanded_element = expand_path_alias(
                                path_alias_sheet, element)
                            expanded_path.append(expanded_element)
                        else:
                            expanded_path.append(element)

                    path_list.append(expanded_path)

                else:
                    path_list.append(f'{section}.{value}')

            if '$loop' in section:

                in_loop = True

                # get csv name and open it
                csv_name = section.strip().split()[-1]
                csv_path = f'./continuous_data/{csv_name}'

                with open(csv_path, 'r') as f:
                    continuous_data = list(csv.reader(f))

                # get loop lines
                loop_start = cell[0].row
                loop_lines = get_loop_lines(
                    sheet, loop_start, path_cell, value_cell)
                expanded_loop_lines = expand_loop_lines(
                    loop_lines, continuous_data, path_alias_sheet, group_alias_sheet)

                [path_list.append(line) for line in expanded_loop_lines]

    return path_list


def expand_group_alias(group_alias_name: str,
                       group_alias_sheet: dict):
    '''
    A function that returns an array containing the lines associated to a group alias

    :param group_alias_name: the name of the group alias
    :param group_alias_sheet: the sheet containing the group aliases
    :return: :list:
    '''
    alias_name_cell = find_cell_location(group_alias_sheet, 'Group Alias Name')
    alias_path_cell = find_cell_location(group_alias_sheet, 'path')

    group_alias_paths = []

    for cell in group_alias_sheet.iter_rows(alias_name_cell[0]+1, group_alias_sheet.max_row, alias_name_cell[1], alias_path_cell[1]+1):
        alias_name = cell[0].value.lower()
        alias_path = cell[-2].value
        alias_value = cell[-1].value

        if alias_name == group_alias_name:
            group_alias_paths.append([alias_path, alias_value])

    return group_alias_paths


def expand_loop_lines(loop_lines: list,
                      continuous_data: list,
                      path_alias_sheet: dict,
                      group_alias_sheet: dict):
    '''
    A function to expand the loop lines and insert the data contained in the csv file
    Returns a list containing the paths and values

    :param loop_lines: the lines contained in the for loop
    :param continuous_data: the data to insert into the loop lines
    :param path_alias_sheet: the sheet containing the path aliases
    :param group_alias_sheet: the group alias sheet
    :return: :list:
    '''
    index = 0
    expanded_loop_lines = []
    for line in continuous_data:
        for arr in loop_lines:
            path = arr[0]
            value = arr[1]

            expanded_line = []

            # expand the path
            if '%' in path:
                path = path.replace('%', str(index))
            if '$$' in path:
                # get the alias name
                group_alias_name = path.split(
                    '.')[-1].replace('$$', '').strip().lower()
                remainder_path = '.'.join(path.split('.')[:-1])

                if '$' in remainder_path:
                    # get alias
                    split_path = remainder_path.split('.')
                    expanded_path = []
                    for element in split_path:
                        if '$' in element:
                            expanded_element = expand_path_alias(
                                path_alias_sheet, element)
                            expanded_path.append(expanded_element)
                        else:
                            expanded_path.append(element)

                    remainder_path = '.'.join(expanded_path)

                # get the expanded list of paths
                group_alias_paths = expand_group_alias(
                    group_alias_name, group_alias_sheet)

                # loop over list and add each line item to path list
                for path in group_alias_paths:
                    if len(remainder_path) == 0:
                        expanded_loop_lines.append(path)
                    else:
                        expanded_loop_lines.append(
                            [f'{remainder_path}.{path[0]}', path[1]])
            elif '$' in path:
                # get alias
                split_path = path.split('.')
                expanded_path = []
                for element in split_path:
                    if '$' in element:
                        expanded_element = expand_path_alias(
                            path_alias_sheet, element)
                        expanded_path.append(expanded_element)
                    else:
                        expanded_path.append(element)

                path = '.'.join(expanded_path)

            # expand the value
            if str(value) != 'None':
                if '$item' in value:
                    item_index = int(value[5:])
                    value = line[item_index]

                expanded_line.append(path)
                expanded_line.append(value)
                expanded_loop_lines.append(expanded_line)
        index += 1

    return expanded_loop_lines


def get_loop_lines(sheet: dict,
                   loop_start: int,
                   path_cell: tuple,
                   value_cell: tuple):
    '''
    A function to collect the path and values contained within a loop

    :param sheet: the sheet conatining the loop
    :param loop_start: the row number the loo
    '''
    loop_lines = []
    for cell in sheet.iter_rows(loop_start+1, sheet.max_row, path_cell[1], value_cell[1]):
        path = translate_cell(cell[0].value)
        value = translate_cell(cell[-1].value)

        if path == '$loopend':
            return loop_lines

        loop_lines.append([path, value])


def expand_path_alias(path_alias_sheet: dict,
                      alias_name: str):
    '''
    A function to return the expanded version on a alias

    :param path_alias_sheet: the sheet containing the path aliases
    :param alias_name: the alias to expand
    '''
    # sort out alias name
    if '[' in alias_name:
        alias_arr = alias_name.split('[')
        alias_name = alias_arr[0]
        alias_index = alias_arr[1][:-1]

    alias_location = find_cell_location(path_alias_sheet, alias_name)
    expanded_alias = path_alias_sheet.cell(
        row=alias_location[0], column=alias_location[1]+1).value

    try:
        expanded_alias = f'{expanded_alias}[{alias_index}]'
        return expanded_alias
    except UnboundLocalError:
        return expanded_alias


def get_whitespace(line: str):
    '''
    A functio to return the amount of white space at the start of a string

    :param line: the line to process
    :return: :int:
    '''
    full_len = len(line)
    stripped = len(line.lstrip())
    whitespace = full_len - stripped

    return whitespace


def get_path(previous_indent: int, indent: int, path: list, stripped: str):
    '''
    A function to get the current data path

    :param previous_indent: the last lines indentation level
    :param indent: the current lines indentation level
    :param path; the previous path
    :param stripped: the data element string stripped of leading and trailing whitespace
    :return: :list:
    '''
    # Get path for A
    if indent == 0:  # No indent
        path = []
        path.append(stripped)
    elif previous_indent == indent:  # Same indent
        path.pop()
        path.append(stripped)
    elif previous_indent < indent:  # More indented
        path.append(stripped)
    elif previous_indent > indent:  # Less indented
        num_less = previous_indent - indent + 1
        path = path[:int(len(path)-num_less)]
        path.append(stripped)
    return path


def get_standard_paths(standard_path: str):
    '''
    This is a function that extracts the paths from a standard in FHIR shorthand format

    :param standard_path: the path to the standard
    :return: :list:
    '''
    # load standard
    wb = openpyxl.load_workbook(standard_path, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    # get row stard and path column
    name_title_cell = find_cell_location(ws, 'Name')

    # get single_indent
    single_indent = get_whitespace(
        str(ws.cell(row=name_title_cell[0]+2, column=name_title_cell[1]).value))

    # make path list
    path_list = []
    path = []
    previous_indent = 0
    for row in ws.iter_rows(name_title_cell[0]+1, ws.max_row, name_title_cell[1], name_title_cell[1]):
        line = str(row[0].value)
        if line != 'None':
            line_indent = int(get_whitespace(line) / single_indent)
            stripped = re.sub(
                re.escape('\\xa0\\xa0\\xa0\\xa0'), r'', line).strip()
            path = get_path(previous_indent, line_indent, path, stripped)

            previous_indent = line_indent
            path_list.append('.'.join(path.copy()).lower().replace(' ', '_'))

    return path_list


def remove_indexing(path: str):
    '''
    A function to remove the indexes from a FHIR shorthand path

    :param path: the FHIR path to process
    :return: :str:
    '''
    path_arr = path.split('.')
    return '.'.join([path.split('[')[0] if '[' in path else path for path in path_arr])


def extract_paths(path_list, plain: bool):
    '''
    A function to retrieve the paths from the path/value list and remove the indexes if required

    :param path_list: the path + value list
    :param plain: boolean value that decides weather to remove the indexes
    :return: :list:
    '''
    new__path_list = []
    for arr in path_list:
        path = arr[0]
        path = remove_indexing(path)

        new__path_list.append(path)

    return new__path_list


def get_paths_and_values(path_list: list):
    '''
    A function to get the paths and values in one list

    :param path_list: the path list to process
    :return: :list:
    '''
    combined_list = []
    for arr in path_list:
        combined = f'{arr[0]}.{str(arr[1])}'
        split = combined.split('.')
        combined_list.append(split)

    return combined_list


def validate_path_list(path_list, standard_path_lists):
    '''
    A function to validate the paths in a path list against the paths in givern standards

    :param path_list: the path list to validate
    :param standard_path_lists: a list of path lists that are generated from the standards
    :return: :list:
    '''
    invalid_paths = []

    for path in path_list:
        found = False
        for standard_path_list in standard_path_lists:
            if not found:
                for standard_path in standard_path_list:
                    if path == standard_path:
                        found = True
                        break
            else:
                break
        if not found:
            invalid_paths.append(path)

    return invalid_paths


def trim_first_elements(path_list: list):
    '''
    A function that removes the first element of each path arr in a path list

    :param path_list: the path list to process
    :return: :list:
    '''
    new_path_list = []
    for path in path_list:
        new_path = path[1:]
        new_path_list.append(new_path)

    return new_path_list


def sort_paths(start_paths: list):
    '''
    A function that sorts the paths in a path list by their first index

    :param start_paths: the unformatted paths
    :return: :list:
    '''
    sorted = []
    for path in start_paths:
        if not sorted:
            sorted.append([path])
        else:
            found = False
            for lists in sorted:
                key = lists[0][0]
                if path[0] == key:
                    lists.append(path)
                    found = True
                    break
            if not found:
                sorted.append([path])
    return sorted


def create_object(start_paths: list):
    '''
    A recursive function that creates a JSON object from a path list

    :param start_paths: the path list
    :return: :dict:
    '''
    result = {}
    my_paths = sort_paths(start_paths)
    for path_lists in my_paths:
        if len(path_lists) == 1 and len(path_lists[0]) == 2:
            matches = re.split(r".*(\[[0-9*]*\]).*", path_lists[0][0])
            if len(matches) != 1:
                try:
                    result[path_lists[0][0].split(
                        "[")[0]].append(path_lists[0][1])
                except:
                    result.update(
                        {path_lists[0][0].split("[")[0]: [path_lists[0][1]]})
            else:
                result.update({path_lists[0][0]: path_lists[0][1]})
        else:
            matches = re.split(r".*(\[[0-9*]*\]).*", path_lists[0][0])
            if len(matches) != 1:
                try:
                    result[path_lists[0][0].split("[")[0]].append(
                        create_object(trim_first_elements(path_lists)))
                except:
                    result.update({path_lists[0][0].split("[")[0]: [
                                  create_object(trim_first_elements(path_lists))]})
            else:
                try:
                    result.update({path_lists[0][0]: create_object(
                        trim_first_elements(path_lists))})
                except:
                    return result
    return result


def rem_new_line(text: str):
    '''
    A function to remove new line charictors from within the middle of text

    :param text: the text to remove it from
    :return: :str:
    '''
    pattern = re.compile(r'(?<=[a-zA-Z0-9.])\s*\n\s*(?=[a-zA-Z0-9.])')
    new_str = re.sub(pattern, ' ', text)

    return new_str


def filter(node, implementationGuidance: bool):
    if isinstance(node, dict):
        retVal = {}
        for key in node:
            if key == 'desc':
                if '#text' in node[key][0]:
                    retVal['description'] = html.unescape(
                        node[key][0]['#text'])
            elif key == 'conformance':
                retVal['mro'] = node[key]
            elif key == 'shortName':
                retVal['name'] = node[key]
            elif key == 'operationalization':
                retVal['valueSets'] = re.sub(
                    r'(?<=release=)[a-z][0-9]*', '', html.unescape(node[key][0]['#text'])).replace('&amp;', '&')
            elif key == 'minimumMultiplicity':
                retVal[key] = node[key]
            elif key == 'maximumMultiplicity':
                retVal[key] = node[key]
            elif key == 'type':
                retVal['type'] = node[key]
            elif key == 'valueDomain':
                if node[key][0]['type'] != 'code':
                    if node[key][0]['type'] != 'ordinal':
                        retVal[key] = node[key].copy()
            elif key == 'context':
                if implementationGuidance:
                    retVal['implementationGuidance'] = rem_new_line(re.sub(
                        r'(?<=release=)[a-z][0-9]*', '', html.unescape(node[key][0]['#text'])).replace('&amp;', '&'))
            elif isinstance(node[key], dict) or isinstance(node[key], list):
                if key not in ['relationship', 'implementation']:
                    child = filter(node[key], implementationGuidance)
                    if child:
                        retVal[key] = child

        if retVal:
            return retVal
        else:
            return None

    elif isinstance(node, list):
        retVal = []
        for entry in node:
            if isinstance(entry, str):
                retVal.append(entry)
            elif isinstance(entry, dict) or isinstance(entry, list):
                child = filter(entry, implementationGuidance)
                if child:
                    retVal.append(child)
        if retVal:
            return retVal
        else:
            return None





def create_false_path_excel(excel_path: str, summaries: list): 
    '''
    A functio that will create and savce an excel documenbt that contains the false paths in an easy toi view way

    :param excel_path: the save path for the excel document
    :param summaries: the false paths for the personae
    '''

    for summary in summaries:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Summary'

        for sheet in summary['sheets']:
            if len(sheet['invalid_paths']) != 0:
                sheet_name = sheet['sheet_name']
                for path in sheet['invalid_paths']:
                    ws.append([sheet_name, path])


        wb.save(f'{excel_path}/{summary["personae"]}_false_paths.xlsx')





    